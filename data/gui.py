import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import queue
from datetime import datetime
import os
import shutil
import pandas as pd
from TikTokLive import TikTokLiveClient
from TikTokLive.events import CommentEvent
from detector import extract_phone_and_suite
from storage import save_phone, TXT_FILE, CSV_FILE
from config import USERNAME, PROXY
# PDF
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch


class TikTokLiveGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("TikTok Live Monitor - Détection de Numéros")
        self.root.geometry("1000x700")
        self.root.configure(bg="#1a1a1a")

        # Variables
        self.client = None
        self.client_thread = None
        self.is_running = False
        self.was_connected = False
        self.username_var = tk.StringVar(value=USERNAME)
        self.status_var = tk.StringVar(value="⏸️ Arrêté")
        self.comments_count = 0
        self.phones_count = 0
        self.message_queue = queue.Queue()

        self.check_queue()
        self.create_widgets()

    def create_widgets(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), background='#1a1a1a', foreground='#ffffff')
        style.configure('Status.TLabel', font=('Segoe UI', 12), background='#1a1a1a', foreground='#00ff00')

        # Couleurs
        bg_dark = "#0d1117"
        bg_medium = "#161b22"
        bg_card = "#21262d"
        accent_pink = "#ff0055"
        accent_green = "#00ff88"
        accent_blue = "#58a6ff"

        self.root.configure(bg=bg_dark)
        main_frame = tk.Frame(self.root, bg=bg_dark)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Header
        header_frame = tk.Frame(main_frame, bg=bg_card, relief=tk.FLAT)
        header_frame.pack(fill=tk.X, pady=(0, 15))
        header_inner = tk.Frame(header_frame, bg=bg_card)
        header_inner.pack(fill=tk.X, padx=20, pady=15)
        title_label = tk.Label(header_inner, text="🔴 TikTok Live Monitor",
                               font=('Segoe UI', 20, 'bold'), bg=bg_card, fg=accent_pink)
        title_label.pack()
        subtitle_label = tk.Label(header_inner, text="Détection automatique de numéros de téléphone",
                                  font=('Segoe UI', 10), bg=bg_card, fg="#8b949e")
        subtitle_label.pack(pady=(5, 0))

        # Contrôles
        control_frame = tk.Frame(main_frame, bg=bg_card, relief=tk.FLAT)
        control_frame.pack(fill=tk.X, pady=(0, 15))

        # Username
        username_frame = tk.Frame(control_frame, bg=bg_card)
        username_frame.pack(side=tk.LEFT, padx=15, pady=15)
        tk.Label(username_frame, text="Username TikTok:", font=('Segoe UI', 10, 'bold'),
                 bg=bg_card, fg="#f0f6fc").pack(side=tk.LEFT, padx=(0, 8))
        username_entry = tk.Entry(
            username_frame,
            textvariable=self.username_var,
            width=25,
            font=('Segoe UI', 10),
            bg=bg_medium,
            fg="#f0f6fc",
            insertbackground="#f0f6fc",
            relief=tk.FLAT,
            bd=5
        )
        username_entry.pack(side=tk.LEFT)

        # Boutons
        button_frame = tk.Frame(control_frame, bg=bg_card)
        button_frame.pack(side=tk.LEFT, padx=20, pady=15)
        self.start_button = tk.Button(button_frame, text="▶ Démarrer",
                                      command=self.start_client, bg="#238636",
                                      fg="white", font=('Segoe UI', 10, 'bold'),
                                      width=14, height=2, relief=tk.FLAT, cursor="hand2",
                                      activebackground="#2ea043", activeforeground="white")
        self.start_button.pack(side=tk.LEFT, padx=5)
        self.stop_button = tk.Button(button_frame, text="⏹ Arrêter",
                                     command=self.stop_client, bg="#da3633",
                                     fg="white", font=('Segoe UI', 10, 'bold'),
                                     width=14, height=2, state=tk.DISABLED, relief=tk.FLAT,
                                     cursor="hand2", activebackground="#f85149", activeforeground="white")
        self.stop_button.pack(side=tk.LEFT, padx=5)
        self.download_button = tk.Button(button_frame, text="💾 Télécharger PDF",
                                         command=self.download_files, bg=accent_blue,
                                         fg="white", font=('Segoe UI', 10, 'bold'),
                                         width=18, height=2, relief=tk.FLAT, cursor="hand2",
                                         activebackground="#79c0ff", activeforeground="white")
        self.download_button.pack(side=tk.LEFT, padx=5)

        # Status
        status_frame = tk.Frame(control_frame, bg=bg_card)
        status_frame.pack(side=tk.RIGHT, padx=20, pady=15)
        tk.Label(status_frame, text="Statut:", font=('Segoe UI', 10, 'bold'),
                 bg=bg_card, fg="#f0f6fc").pack(side=tk.LEFT, padx=(0, 8))
        self.status_label = tk.Label(status_frame, textvariable=self.status_var,
                                     font=('Segoe UI', 10, 'bold'), bg=bg_card, fg=accent_green)
        self.status_label.pack(side=tk.LEFT)

        # Stats
        stats_frame = tk.Frame(main_frame, bg=bg_card, relief=tk.FLAT)
        stats_frame.pack(fill=tk.X, pady=(0, 15))
        stats_inner = tk.Frame(stats_frame, bg=bg_card)
        stats_inner.pack(fill=tk.X, padx=20, pady=12)
        self.comments_label = tk.Label(stats_inner, text="💬 Commentaires: 0",
                                       font=('Segoe UI', 12, 'bold'), bg=bg_card, fg="#f0f6fc")
        self.comments_label.pack(side=tk.LEFT, padx=30)
        separator = tk.Frame(stats_inner, bg="#30363d", width=2, height=25)
        separator.pack(side=tk.LEFT, padx=20)
        self.phones_label = tk.Label(stats_inner, text="📞 Numéros détectés: 0",
                                     font=('Segoe UI', 12, 'bold'), bg=bg_card, fg=accent_green)
        self.phones_label.pack(side=tk.LEFT, padx=30)

        # Contenu
        content_frame = tk.Frame(main_frame, bg=bg_dark)
        content_frame.pack(fill=tk.BOTH, expand=True)
        left_frame = tk.Frame(content_frame, bg=bg_card, relief=tk.FLAT)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 7))
        left_header = tk.Frame(left_frame, bg=bg_card)
        left_header.pack(fill=tk.X, pady=(12, 8), padx=12)
        tk.Label(left_header, text="💬 Commentaires en temps réel",
                 font=('Segoe UI', 13, 'bold'), bg=bg_card, fg="#f0f6fc").pack()
        self.comments_text = scrolledtext.ScrolledText(
            left_frame, height=20, width=50, bg=bg_medium, fg="#c9d1d9",
            font=('Consolas', 10), wrap=tk.WORD, relief=tk.FLAT, bd=8,
            insertbackground="#f0f6fc"
        )
        self.comments_text.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))

        right_frame = tk.Frame(content_frame, bg=bg_card, relief=tk.FLAT)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(7, 0))
        right_header = tk.Frame(right_frame, bg=bg_card)
        right_header.pack(fill=tk.X, pady=(12, 8), padx=12)
        tk.Label(right_header, text="📞 Numéros de téléphone détectés",
                 font=('Segoe UI', 13, 'bold'), bg=bg_card, fg=accent_green).pack()
        self.phones_text = scrolledtext.ScrolledText(
            right_frame, height=20, width=50, bg=bg_medium, fg=accent_green,
            font=('Consolas', 10), wrap=tk.WORD, relief=tk.FLAT, bd=8,
            insertbackground=accent_green
        )
        self.phones_text.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))

    # --- Gestion queue ---
    def check_queue(self):
        try:
            while True:
                msg_type, data = self.message_queue.get_nowait()
                if msg_type == "comment":
                    self.add_comment(data['user'], data['text'])
                elif msg_type == "phone":
                    self.add_phone(data['phone'], data['suite'], data['user'])
                elif msg_type == "status":
                    self.status_var.set(data)
                elif msg_type == "error":
                    self.add_error(data)
                elif msg_type == "live_ended":
                    self.handle_live_ended()
        except queue.Empty:
            pass
        finally:
            self.root.after(100, self.check_queue)

    # --- Ajout texte ---
    def add_comment(self, user, text):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.comments_text.insert(tk.END, f"[{timestamp}] {user}: {text}\n")
        self.comments_text.see(tk.END)
        self.comments_count += 1
        self.comments_label.config(text=f"💬 Commentaires: {self.comments_count}")

    def add_phone(self, phone, suite, user):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.phones_text.insert(tk.END, f"[{timestamp}] 📞 {phone}\n")
        self.phones_text.insert(tk.END, f"    👤 {user}\n")
        if suite.strip():
            self.phones_text.insert(tk.END, f"    📝 {suite}\n")
        self.phones_text.insert(tk.END, f"{'-'*50}\n")
        self.phones_text.see(tk.END)
        self.phones_count += 1
        self.phones_label.config(text=f"📞 Numéros détectés: {self.phones_count}")

    def add_error(self, error_msg):
        self.comments_text.insert(tk.END, f"⚠️ ERREUR: {error_msg}\n", "error")
        self.comments_text.see(tk.END)
        self.comments_text.tag_config("error", foreground="#ff0000")

    # --- TikTok Client ---
    def create_client(self):
        username = self.username_var.get().strip()
        if not username:
            messagebox.showerror("Erreur", "Veuillez entrer un username TikTok")
            return None

        client_config = {"unique_id": username}
        if PROXY:
            client_config["proxy"] = PROXY
        client = TikTokLiveClient(**client_config)

        @client.on(CommentEvent)
        async def on_comment(event: CommentEvent):
            try:
                text = event.comment
                user = event.user.nick_name
                self.message_queue.put(("comment", {"user": user, "text": text}))
                results = extract_phone_and_suite(text)
                for phone, suite in results:
                    if save_phone(phone, suite, user):
                        self.message_queue.put(("phone", {"phone": phone, "suite": suite, "user": user}))
            except Exception as e:
                self.message_queue.put(("error", str(e)))

        return client

    def run_client(self):
        try:
            self.message_queue.put(("status", "🟢 Connexion en cours..."))
            self.client = self.create_client()
            if not self.client:
                self.is_running = False
                self.message_queue.put(("status", "⏸️ Arrêté"))
                return
            self.message_queue.put(("status", "🟢 Connecté - En attente du LIVE..."))
            self.was_connected = True
            self.client.run()
        except Exception as e:
            self.message_queue.put(("error", str(e)))
            self.is_running = False
            self.root.after(0, self.update_buttons)

    def start_client(self):
        if self.is_running: return
        self.is_running = True
        self.was_connected = False
        self.update_buttons()
        self.client_thread = threading.Thread(target=self.run_client, daemon=True)
        self.client_thread.start()

    def stop_client(self):
        if not self.is_running: return
        self.is_running = False
        self.was_connected = False
        self.message_queue.put(("status", "⏸️ Arrêt en cours..."))
        if self.client:
            try:
                self.client.stop()
            except:
                pass
        self.message_queue.put(("status", "⏸️ Arrêté"))
        self.update_buttons()

    def handle_live_ended(self):
        self.is_running = False
        self.update_buttons()
        self.message_queue.put(("status", "🔴 Live terminé"))
        self.root.after(500, self.ask_download_after_live)

    def ask_download_after_live(self):
        if messagebox.askyesno("Live terminé",
                               f"Le live est terminé.\n\n"
                               f"💬 Commentaires reçus: {self.comments_count}\n"
                               f"📞 Numéros détectés: {self.phones_count}\n\n"
                               f"Voulez-vous télécharger les fichiers des commentaires extraits ?"):
            self.download_files()

    # --- Export fichiers ---
    def download_files(self):
        try:
            if not os.path.exists(CSV_FILE):
                messagebox.showwarning("Aucun fichier",
                                       "Aucun fichier de commentaires trouvé.\n"
                                       "Les fichiers seront créés lors de la détection de numéros.")
                return

            dest_folder = filedialog.askdirectory(title="Choisir le dossier de destination pour le PDF")
            if not dest_folder:
                return

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            username = self.username_var.get().strip() or "tiktok"
            folder_name = f"tiktok_live_{username}_{timestamp}"
            dest_path = os.path.join(dest_folder, folder_name)
            os.makedirs(dest_path, exist_ok=True)
            files_copied = []

            if os.path.exists(CSV_FILE):
                # Générer PDF uniquement
                try:
                    pdf_file = os.path.join(dest_path, "phones.pdf")
                    self.create_pdf_from_csv(CSV_FILE, pdf_file)
                    if os.path.exists(pdf_file):
                        files_copied.append("phones.pdf")
                except Exception as e:
                    print(f"Erreur lors de la création du fichier PDF: {e}")

            if files_copied:
                messagebox.showinfo("Téléchargement réussi",
                                    f"Fichiers téléchargés avec succès !\n\n"
                                    f"📁 Dossier: {dest_path}\n\n"
                                    f"Fichiers générés:\n" + "\n".join(f"  • {f}" for f in files_copied))
            else:
                messagebox.showwarning("Aucun fichier", "Aucun fichier à télécharger.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors du téléchargement:\n{str(e)}")

    def create_excel_from_csv(self, csv_file, excel_file):
        """Crée un fichier Excel avec formatage depuis le CSV"""
        try:
            # Lire le CSV
            df = pd.read_csv(csv_file, encoding="utf-8")
            
            if df.empty:
                return
            
            # Gérer les colonnes manquantes
            if 'Time' not in df.columns:
                df['Time'] = 'N/A'
            if 'User' not in df.columns:
                df['User'] = ''
            if 'Phone' not in df.columns:
                df['Phone'] = ''
            if 'Suite_Commentaire' not in df.columns:
                df['Suite_Commentaire'] = ''
            
            # Créer un nouveau DataFrame avec les colonnes dans le bon ordre
            # Chaque colonne est extraite séparément pour garantir la séparation
            result_df = pd.DataFrame({
                'Time': df['Time'].astype(str).replace('nan', 'N/A').fillna('N/A'),
                'User': df['User'].astype(str).replace('nan', '').fillna(''),
                'Phone': df['Phone'].astype(str).replace('nan', '').fillna(''),
                'Suite Commentaire': df['Suite_Commentaire'].astype(str).replace('nan', '').fillna('')
            })
            
            # Créer le fichier Excel
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Numéros Détectés')
                
                worksheet = writer.sheets['Numéros Détectés']
                
                # Import des styles
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Style de l'en-tête
                header_font = Font(bold=True, color="FFFFFF", size=12)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                border_style = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Appliquer le formatage à l'en-tête
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_style
                
                # Ajuster la largeur des colonnes
                worksheet.column_dimensions['A'].width = 20  # Time
                worksheet.column_dimensions['B'].width = 25  # User
                worksheet.column_dimensions['C'].width = 15  # Phone
                worksheet.column_dimensions['D'].width = 40  # Suite Commentaire
                
                # Appliquer le formatage aux données
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.border = border_style
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        if cell.row % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Geler la première ligne
                worksheet.freeze_panes = 'A2'
                
        except Exception as e:
            raise Exception(f"Erreur Excel: {str(e)}")

    def create_pdf_from_csv(self, csv_file, pdf_file):
        """Crée un PDF formaté avec Time, User, Phone, Suite Commentaire"""
        try:
            if not os.path.exists(csv_file):
                return
            
            # Lire le CSV avec gestion des en-têtes manquants
            # Forcer la colonne Phone à être lue comme string pour éviter les .0
            expected_cols = ["Time", "User", "Phone", "Suite_Commentaire"]
            dtype_dict = {'Phone': str}
            
            # D'abord essayer de lire avec en-têtes
            try:
                df = pd.read_csv(csv_file, encoding="utf-8", dtype=dtype_dict)
                # Vérifier si les colonnes attendues existent
                if not all(col in df.columns for col in expected_cols):
                    # Les colonnes attendues ne sont pas là, relire sans en-têtes
                    df = pd.read_csv(csv_file, encoding="utf-8", header=None, 
                                    names=expected_cols, dtype=dtype_dict)
            except Exception as e:
                # Si erreur, essayer sans en-têtes
                try:
                    df = pd.read_csv(csv_file, encoding="utf-8", header=None, 
                                    names=expected_cols, dtype=dtype_dict)
                except:
                    raise Exception(f"Impossible de lire le fichier CSV: {e}")
            
            if df.empty:
                return
            
            # Gérer les colonnes manquantes
            if 'Time' not in df.columns:
                df['Time'] = 'N/A'
            if 'User' not in df.columns:
                df['User'] = ''
            if 'Phone' not in df.columns:
                df['Phone'] = ''
            if 'Suite_Commentaire' not in df.columns:
                df['Suite_Commentaire'] = ''
            
            # Créer le document PDF
            doc = SimpleDocTemplate(pdf_file, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Titre
            title_style = styles['Heading1']
            title_style.alignment = 1  # Centré
            elements.append(Paragraph("📞 Numéros de Téléphone Détectés", title_style))
            elements.append(Spacer(1, 12))
            
            # Date de génération
            date_style = styles['Normal']
            date_text = Paragraph(f"<i>Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}</i>", date_style)
            elements.append(date_text)
            elements.append(Spacer(1, 20))
            
            # Préparer les données pour le tableau
            data = [["Time", "User", "Phone", "Suite Commentaire"]]
            
            for _, row in df.iterrows():
                time_val = str(row['Time'])[:19] if pd.notna(row.get('Time', '')) else 'N/A'
                user_val = str(row['User']) if pd.notna(row.get('User', '')) else ''
                
                # Convertir le téléphone en string et retirer le .0 si présent (problème float)
                phone_raw = row.get('Phone', '')
                if pd.notna(phone_raw):
                    phone_str = str(phone_raw)
                    # Retirer .0 à la fin si c'est un nombre entier (cas où pandas lit comme float)
                    if '.' in phone_str and phone_str.replace('.0', '').replace('-', '').isdigit():
                        phone_val = phone_str.rstrip('0').rstrip('.')
                    else:
                        phone_val = phone_str
                else:
                    phone_val = ''
                
                suite_val = str(row['Suite_Commentaire']) if pd.notna(row.get('Suite_Commentaire', '')) else ''
                
                data.append([time_val, user_val, phone_val, suite_val])
            
            # Créer le tableau
            table = Table(data, colWidths=[1.5*inch, 1.5*inch, 1.2*inch, 2.8*inch])
            
            # Style du tableau
            table.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2d4a87")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                
                # Données
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
                
                # Bordures
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                
                # Padding
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ]))
            
            elements.append(table)
            doc.build(elements)
            
        except Exception as e:
            raise Exception(f"Erreur PDF: {str(e)}")

    # --- Boutons ---
    def update_buttons(self):
        if self.is_running:
            self.start_button.config(state=tk.DISABLED)
            self.stop_button.config(state=tk.NORMAL)
        else:
            self.start_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)


def main():
    root = tk.Tk()
    app = TikTokLiveGUI(root)
    root.protocol("WM_DELETE_WINDOW", lambda: (app.stop_client(), root.destroy()))
    root.mainloop()


if __name__ == "__main__":
    main()
